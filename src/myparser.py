import openpyxl
import datetime
from os import path




def time2Str(tm):
    if isinstance(tm, datetime.datetime):
        return tm.strftime("%Y/%m/%d %H:%M:%S")
    return tm

def writeExcel(data):
    wb = openpyxl.Workbook()
    new_sheet = wb.active
    new_sheet.title = "数据结果"
    for row in data:
        # print(f"new_row: {row}")
        new_sheet.append([
            row.get("name"),
            time2Str(row.get("last_pay_time")),
            time2Str(row.get("leave_time")),
            row.get("total"),
            row.get("is_cancel"),
            row.get("is_promote"),
            row.get("last_level"),
            row.get("cur_level")
        ])
    current_file_path = path.realpath(__file__)
    current_dir = path.dirname(current_file_path)
    wb.save(path.join(current_dir, '..', 'output', 'output.xlsx'))


def extract_cell_data(cell):
    return cell.value


def extract_data(row, index):
    if index < 8 or row[1] is None or row[1].value is None:
        return
    # cell_type = type(row[20])
    # print(f"type: {cell_type} {row[20]} {row[20].value}")
    return {
        "name": extract_cell_data(row[1]),
        "last_pay_time": extract_cell_data(row[11]),
        "leave_time":extract_cell_data(row[12]),
        "total":extract_cell_data(row[20]),
        "is_cancel":extract_cell_data(row[21]),
        "is_promote":extract_cell_data(row[22]),
        "last_level":extract_cell_data(row[23]),
        "cur_level":extract_cell_data(row[24])
    }




def parseExcel(excelPath, func):
        # 打开工作簿和选定的工作表
    data_list = []
    workbook = openpyxl.load_workbook(excelPath, data_only=True)

    sheet_names = workbook.sheetnames
    first_sheet_name = sheet_names[0]  # 获取第一个工作表的名称
    sheet = workbook[first_sheet_name]
    # 遍历工作表的每一行
    row_index = 0
    for row in sheet.iter_rows():  # 如果你希望获取单元格对象而不是值，移除values_only=True
        # data = {}
        # # for index,cell in enumerate(row):

        # for cell in row:
        #     if(cell is not None):
        #         func(cell, 0, data)
                # print(cell.value)  # 打印整行的数据作为一个元组
        new_data = func(row, row_index)
        if new_data is not None:
            data_list.append(new_data)
        row_index = row_index + 1

    # # 读取特定的单元格
    # cell_value = sheet['A1'].value  # 读取A1单元格的值
    # print(cell_value)

    # 关闭工作簿
    workbook.close()
    return data_list